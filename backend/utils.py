import re
from typing import Optional
from openpyxl import load_workbook
import csv, io

_UNITS_RE = re.compile(
    r"""^\s*
        (?:мл|ml|г|мг|mg|кг|л|l|литр(?:а|ов)?|лит|таб(?:лет)?|капс(?:ул)?|caps?|шт|см|мм)
        \b
    """,
    re.IGNORECASE | re.VERBOSE
)

def _looks_like_unit_after(s: str) -> bool:
    """Проверяем, что сразу после числа идёт обозначение единиц (объём/вес/шт/длина)."""
    # смотрим короткий «хвост» после числа — достаточно первых ~8 символов
    tail = s[:8]
    return _UNITS_RE.match(tail) is not None

def extract_max_relevant_number(title: Optional[str]) -> Optional[int]:
    """
    Достаём МАКСИМАЛЬНОЕ релевантное целое число из title.
    Фильтры:
      - если сразу после числа идёт ед.изм. (мл, г, мг, л, шт и т.п.) — пропускаем число;
      - очень большие числа (>= 100000) считаем кодами — пропускаем;
    """
    if not title:
        return None
    text = str(title)

    candidates = []
    for m in re.finditer(r"(?<!\d)(\d{1,6})(?!\d)", text):
        num_str = m.group(1)
        try:
            num = int(num_str)
        except Exception:
            continue

        # отбрасываем явные «коды»
        if num >= 10000:
            continue

        # если сразу после числа идёт ед.изм. — считаем это объёмом/весом и пропускаем
        tail_start = m.end()
        tail = text[tail_start: tail_start + 8]
        if _looks_like_unit_after(tail):
            continue

        candidates.append(num)

    if candidates:
        return max(candidates)

    # Если хочешь: мягкий фолбэк — всё равно взять наибольшее число
    # (закомментировано, чтобы не увеличивать риск ложных срабатываний)
    #
    # all_nums = [int(m.group(1)) for m in re.finditer(r"(?<!\d)(\d{1,6})(?!\d)", text)]
    # if all_nums:
    #     return max(all_nums)

    return None

def calc_rrc_from_title(title: Optional[str]) -> Optional[float]:
    """
    Правило: если число < 700 -> РРЦ = 1300, иначе РРЦ = 1500.
    Если число не найдено — None (РРЦ не трогаем).
    """
    n = extract_max_relevant_number(title)
    if n is None:
        return None
    if n <= 0:
        return None
    elif n > 800:
        return None
    elif n < 700:
        return 1300.0
    else:
        return 1500.0

# --- ВСПОМОГАЮЩИЕ ФУНКЦИИ ДЛЯ XLSX --- #

HEADER_SELLER_COL_KEYS = [
    "артикул", "nm", "nm id", "nm_id", "код товара", "sku", "artikul"
]
HEADER_RRC_COL_KEYS = [
    "ррц", "рекоменд", "минимальная цена", "минимальная стоимость",
    "recommended", "min price", "minprice"
]

def _norm_header(v):
    if v is None:
        return ""
    s = str(v).strip().lower()
    # убираем лишние пробелы
    s = re.sub(r"\s+", " ", s)
    return s

def _looks_like_nm_id(val):
    """
    nm_id — обычно целое с 6..12 цифрами. Разрешим до 12, чтобы не мешали длинные коды.
    """
    if val is None:
        return False
    s = str(val).strip()
    # иногда из Excel прилетает float или экспоненциальная запись
    # попробуем привести
    try:
        # если строка, вытащим только цифры
        if isinstance(val, str):
            digits = re.sub(r"\D+", "", val)
            if not digits:
                return False
            # 6..12 цифр
            return 6 <= len(digits) <= 12
        elif isinstance(val, int):
            ival = int(val)
            return 100000 <= ival <= 999999999999
        elif isinstance(val, float):
            ival = int(val)
            return 100000 <= ival <= 999999999999
    except Exception:
        return False
    return False

def _parse_nm_id(val):
    """
    Аккуратно парсим nm_id: берём цифры, преобразуем в int, проверяем границы.
    Возвращаем int или None.
    """
    if val is None:
        return None
    try:
        if isinstance(val, (int,)):
            maybe = int(val)
        elif isinstance(val, float):
            maybe = int(val)
        else:
            s = str(val)
            digits = re.sub(r"\D+", "", s)
            if not digits:
                return None
            maybe = int(digits)
        if 100000 <= maybe <= 999999999999:
            return maybe
    except Exception:
        return None
    return None

def _parse_price_like(val):
    """
    Парсим цену вида '1 500', '1500,00', 1500.0 -> float.
    Возвращаем float или None.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    # заменим запятую на точку, уберём пробелы
    s = s.replace(",", ".")
    s = re.sub(r"[^\d.]+", "", s)
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None

def _detect_columns(ws, sample_rows=200):
    """
    Возвращает (idx_nm, idx_rrc) — индексы с 0 (или None).
    Алгоритм:
      A) скорим заголовки (первая строка) — точное «артикул» получает максимальный вес;
      B) если несколько кандидатов — добиваем выбор по данным (голоса _looks_like_nm_id и {1300,1500});
      C) в самом крайнем случае берём колонку с макс. числом «похожих на nm_id» / числовых значений.
    """

    try:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return None, None

    headers_raw = list(first_row or [])
    headers = [_norm_header(h) for h in headers_raw]
    ncols = len(headers)

    # --- скоринг заголовков
    def score_nm_header(h: str) -> int:
        # 100 — точное "артикул"
        if h == "артикул":
            return 100
        # 90 — строка начинаетcя с "артикул"
        if h.startswith("артикул"):
            return 90
        # 80 — содержит слово "nm" / "nm id" / "nm_id"
        if any(k in h for k in ["nm id", "nm_id", " nm ", " nm"]):
            return 80
        # 70 — содержит слова из нашего набора
        if any(k in h for k in HEADER_SELLER_COL_KEYS):
            return 70
        return 0

    def score_rrc_header(h: str) -> int:
        # 100 — точное "ррц"
        if h == "ррц":
            return 100
        # 90 — начинается с "ррц"
        if h.startswith("ррц"):
            return 90
        # 85 — содержит "рекоменд"
        if "рекоменд" in h:
            return 85
        # 80 — содержит "min price"/"минимальн"
        if any(k in h for k in ["минимальн", "min price", "minprice"]):
            return 80
        # 70 — по общему списку ключей
        if any(k in h for k in HEADER_RRC_COL_KEYS):
            return 70
        return 0

    nm_scores = [score_nm_header(h) for h in headers]
    rrc_scores = [score_rrc_header(h) for h in headers]

    # --- сбор голосов по данным (до sample_rows)
    nm_votes = [0] * ncols
    rrc_votes = [0] * ncols
    rrc_1300_1500_votes = [0] * ncols

    for row in ws.iter_rows(min_row=2, max_row=1 + sample_rows, values_only=True):
        for i in range(ncols):
            cell = row[i] if i < len(row) else None
            if _looks_like_nm_id(cell):
                nm_votes[i] += 1
            price = _parse_price_like(cell)
            if price is not None:
                rrc_votes[i] += 1
                if int(round(price)) in (1300, 1500):
                    rrc_1300_1500_votes[i] += 1

    # --- выбор индекса артикулов
    idx_nm = None
    if any(nm_scores):
        best_score = max(nm_scores)
        cand_idxs = [i for i, s in enumerate(nm_scores) if s == best_score]
        if len(cand_idxs) == 1:
            idx_nm = cand_idxs[0]
        else:
            # тай-брейк по голосам данных
            idx_nm = max(cand_idxs, key=lambda i: (nm_votes[i], -i))

    # если по заголовкам не нашли — по данным
    if idx_nm is None and any(nm_votes):
        idx_nm = int(max(range(ncols), key=lambda i: nm_votes[i]))

    # --- выбор индекса РРЦ
    idx_rrc = None
    if any(rrc_scores):
        best_score = max(rrc_scores)
        cand_idxs = [i for i, s in enumerate(rrc_scores) if s == best_score]
        if len(cand_idxs) == 1:
            idx_rrc = cand_idxs[0]
        else:
            # сначала по совпадениям с {1300,1500}, затем просто по числовым
            idx_rrc = max(cand_idxs, key=lambda i: (rrc_1300_1500_votes[i], rrc_votes[i], -i))

    if idx_rrc is None:
        if any(rrc_1300_1500_votes):
            idx_rrc = int(max(range(ncols), key=lambda i: rrc_1300_1500_votes[i]))
        elif any(rrc_votes):
            idx_rrc = int(max(range(ncols), key=lambda i: rrc_votes[i]))

    return idx_nm, idx_rrc

def _norm(s):
    return (str(s).strip() if s is not None else "")

def iter_ozon_csv_rows(file_bytes: bytes):
    """
    Потоковый парсер CSV (оба формата).
    Находит ключевые колонки по русским заголовкам.
    Возвращает словари:
      { 'nm_id': int, 'title': str,
        'price_after': float|None, 'price_before': float|None,
        'seller_id': int|None, 'seller_name': str|None }
    """
    # 1) определим кодировку
    for enc in ("utf-8", "utf-8-sig", "cp1251", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except Exception:
            continue
    else:
        text = file_bytes.decode("utf-8", errors="ignore")

    # 2) читаем CSV (разделитель в примерах — запятая)
    reader = csv.DictReader(io.StringIO(text), delimiter=",")

    # 3) нормализуем имена колонок (в нижний регистр, убираем лишние пробелы)
    def norm_header(h): 
        return _norm(h).lower()

    headers = {norm_header(h): h for h in reader.fieldnames or []}

    # Набор возможных ключей
    key_nm      = next((headers[k] for k in headers if "артикул" in k), None)
    key_title   = next((headers[k] for k in headers if "название" in k), None)
    key_price   = next((headers[k] for k in headers if k == "цена"), None)
    key_price_sp= next((headers[k] for k in headers if "спец" in k and "цена" in k), None)
    key_old     = next((headers[k] for k in headers if "старая цена" in k), None)
    key_seller  = next((headers[k] for k in headers if "продавец" == k), None)
    key_seller_id = next((headers[k] for k in headers if "id продавца" == k), None)

    for row in reader:
        nm_raw = row.get(key_nm) if key_nm else None
        nm_id = _parse_nm_id(nm_raw)
        if not nm_id:
            continue  # пропускаем странные строки/итоги

        title = _norm(row.get(key_title)) if key_title else ""

        # price_after: сначала "Спец. цена", затем "Цена"
        p_after = None
        if key_price_sp:
            p_after = _parse_price_like(row.get(key_price_sp))
        if p_after is None and key_price:
            p_after = _parse_price_like(row.get(key_price))

        # price_before: "Старая цена" (если нет — None/0)
        p_before = _parse_price_like(row.get(key_old)) if key_old else None

        # продавец (во втором формате есть, в первом — нет)
        seller_name = _norm(row.get(key_seller)) if key_seller else None
        sid_raw = row.get(key_seller_id) if key_seller_id else None
        try:
            seller_id = int(str(sid_raw).strip()) if sid_raw not in (None, "", "nan") else None
        except Exception:
            seller_id = None

        yield {
            "nm_id": nm_id,
            "title": title,
            "price_after": float(p_after) if p_after is not None else None,
            "price_before": float(p_before) if p_before is not None else None,
            "seller_id": seller_id,
            "seller_name": (seller_name or None),
        }