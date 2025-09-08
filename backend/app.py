from flask import Flask, request, jsonify
from flask_cors import CORS
from wb_parser import fetch_wb_price
from db import (
    list_products as db_list,
    upsert_product,
    delete_product as db_delete,
    set_rrc,
    list_nm_ids_for_refresh,
    count_needing_refresh,
    list_sellers_with_violations,
    delete_all_products,
    acquire_advisory_lock, release_advisory_lock,
    count_all_rows, count_violations,
    list_products_page, list_products_page_violations,
    list_violations_for_seller,
    sales_24h_for_nm_list
)
from utils import calc_rrc_from_title, _parse_nm_id, _parse_price_like, _detect_columns, iter_ozon_csv_rows
from telegram_client import send_violation_alert

import os
from math import floor
from typing import Optional
from dotenv import load_dotenv
from openpyxl import load_workbook
import time

load_dotenv()  # загрузим .env заранее
DB_TABLE_DEFAULT = os.getenv("DB_TABLE", "products")
DB_TABLE_OZON = os.getenv("DB_TABLE_OZON", "products_ozon")

WORK_START_HOUR = int(os.getenv("WORK_START_HOUR", "9"))
WORK_END_HOUR   = int(os.getenv("WORK_END_HOUR", "20"))
WORK_TZ_OFFSET  = int(os.getenv("WORK_TZ_OFFSET", "0"))  # смещение в часах от UTC

# ====== Фиолетовая цена (UI) ======
UI_WALLET_PERCENT = float(os.getenv("UI_WALLET_PERCENT", "0"))
UI_ROUND_THRESHOLD = float(os.getenv("UI_ROUND_THRESHOLD", "0"))
UI_ROUND_STEP = float(os.getenv("UI_ROUND_STEP", "1"))

# ====== Пакетное обновление ======
BATCH_REFRESH_LIMIT = int(os.getenv("BATCH_REFRESH_LIMIT", "20"))

# ====== Загрузка файлов (xlsx) ======
UPLOAD_ALLOWED_EXT = {".xlsx"}

# ====== Разрешённые таблицы ======
ALLOWED_TABLES = {DB_TABLE_DEFAULT, DB_TABLE_OZON}

def resolve_table() -> str:
    """
    Определяем целевую таблицу из query-параметра ?table=...
    По умолчанию — DB_TABLE_DEFAULT.
    Белый список: products, products_ozon (можно расширить переменными окружения).
    """
    t = (request.args.get("table") or "").strip()
    if t in ALLOWED_TABLES:
        return t
    # допускаем «чистые» имена, но всё равно проверяем whitelist — безопасность SQL
    return DB_TABLE_DEFAULT

def now_local_hour() -> int:
    t = time.time() + WORK_TZ_OFFSET * 3600
    return time.gmtime(t).tm_hour

def is_work_time() -> bool:
    h = now_local_hour()
    if WORK_START_HOUR <= WORK_END_HOUR:
        return WORK_START_HOUR <= h < WORK_END_HOUR
    return h >= WORK_START_HOUR or h < WORK_END_HOUR

def calc_ui_price_from_product(base_price: Optional[float]) -> Optional[int]:
    """Считаем UI-цену (рубли, целое) по заданным правилам округления."""
    if base_price is None:
        return None
    try:
        base = float(base_price)
        if base <= 0:
            return None
        p = UI_WALLET_PERCENT
        thr = UI_ROUND_THRESHOLD
        step = UI_ROUND_STEP if UI_ROUND_STEP > 0 else 1.0

        raw = base * (1.0 - p) if p > 0 else base
        if thr > 0 and base >= thr:
            ui = floor(raw / step) * step
        else:
            ui = raw
        return int(floor(ui))
    except Exception:
        return None

def _infer_unavailable(data: dict) -> bool:
    """Эвристика недоступности товара: нет цен и/или флаг in_stock/available=false."""
    for k in ("in_stock", "available"):
        if k in data and data.get(k) is False:
            return True
    pa = float(data.get("price_after_seller_discount") or 0)
    pb = float(data.get("price_before_discount") or 0)
    return (pa <= 0 and pb <= 0)

def _read_html_from_request() -> str:
    # 1) multipart: поля 'html' или 'file'
    f = request.files.get("html") or request.files.get("file")
    data = f.read() if f else (request.get_data() or b"")
    # 2) поддержка gzip
    if (request.headers.get("Content-Encoding") or "").lower() == "gzip":
        try:
            import gzip
            data = gzip.decompress(data)
        except Exception:
            pass
    # 3) декодирование
    for enc in ("utf-8", "cp1251", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")

application = Flask(__name__)
CORS(
    application,
    resources={r"/api/*": {"origins": "*"}},
    methods=["GET", "POST", "DELETE", "PATCH", "OPTIONS"],
)

@application.get("/api/health")
def health():
    return jsonify({"ok": True})

@application.get("/api/products")
def list_products():
    """
    Пагинация:
      ?limit=100&offset=0&only_violations=0|1&sort=nm_id|seller_name|price_after_seller_discount|updated_at|rrc
      &dir=asc|desc
      &table=products|products_ozon
    """
    try:
        table = resolve_table()

        limit = request.args.get("limit", default=100, type=int)
        offset = request.args.get("offset", default=0, type=int)
        only_viol = request.args.get("only_violations", default=0, type=int)
        sort = (request.args.get("sort") or "nm_id").strip()
        dir_ = (request.args.get("dir") or "asc").strip().lower()

        if limit < 1: limit = 100
        if limit > 500: limit = 500
        if offset < 0: offset = 0
        if dir_ not in ("asc", "desc"): dir_ = "asc"

        sort_map = {
            "nm_id": "nm_id",
            "seller_name": "seller_name",
            "price_after_seller_discount": "price_after_seller_discount",
            "updated_at": "updated_at",
            "rrc": "rrc",
        }
        sort_col = sort_map.get(sort, "nm_id")

        if only_viol:
            total_rows = count_violations(table=table)
            items = list_products_page_violations(limit=limit, offset=offset, order_by=sort_col, order_dir=dir_, table=table)
        else:
            total_rows = count_all_rows(table=table)
            items = list_products_page(limit=limit, offset=offset, order_by=sort_col, order_dir=dir_, table=table)
        
        nm_ids = [int(it["nm_id"]) for it in items]  # текущая страница
        try:
            sold_map = sales_24h_for_nm_list(nm_ids)
        except Exception:
            sold_map = {}

        for it in items:
            it["sold_24h"] = int(sold_map.get(int(it["nm_id"]), 0))

        page = (offset // limit) + 1
        total_pages = (total_rows + limit - 1) // limit

        return jsonify({
            "items": items,
            "limit": limit,
            "offset": offset,
            "page": page,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "only_violations": 1 if only_viol else 0,
            "sort": sort,
            "dir": dir_,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.post("/api/products")
def add_product():
    table = resolve_table()
    body = request.get_json(silent=True) or {}
    nm_id = body.get("nm_id")
    try:
        nm_id = int(nm_id)
    except Exception:
        return jsonify({"ok": False, "error": "nm_id должен быть числом"}), 400
    try:
        data = fetch_wb_price(nm_id)
        unavail = _infer_unavailable(data)
        rrc_auto = calc_rrc_from_title(data.get("title"))

        if unavail:
            upsert_product(
                nm_id=data["nm_id"],
                brand=data.get("brand", "") or "",
                title=data.get("title", "") or "",
                price_before=0.0,
                price_after=-1.0,
                seller_id=data.get("seller_id"),
                seller_name=data.get("seller_name") or None,
                ui_price=None,
                table=table,
            )
        else:
            ui_price = calc_ui_price_from_product(float(data.get("price_after_seller_discount") or 0))
            upsert_product(
                nm_id=data["nm_id"],
                brand=data.get("brand", "") or "",
                title=data.get("title", "") or "",
                price_before=float(data.get("price_before_discount") or 0),
                price_after=float(data.get("price_after_seller_discount") or 0),
                seller_id=data.get("seller_id"),
                seller_name=data.get("seller_name") or None,
                ui_price=int(ui_price) if ui_price is not None else None,
                table=table,
            )

        if rrc_auto is not None:
            set_rrc(data["nm_id"], rrc_auto, table=table)

        return jsonify({"ok": True, "item": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.post("/api/products/<int:nm_id>/refresh")
def refresh_product(nm_id: int):
    table = resolve_table()
    try:
        data = fetch_wb_price(nm_id)
        unavail = _infer_unavailable(data)
        rrc_auto = calc_rrc_from_title(data.get("title"))

        if unavail:
            upsert_product(
                nm_id=data["nm_id"],
                brand=data.get("brand", "") or "",
                title=data.get("title", "") or "",
                price_before=0.0,
                price_after=-1.0,
                seller_id=data.get("seller_id"),
                seller_name=data.get("seller_name") or None,
                ui_price=None,
                table=table,
            )
        else:
            ui_price = calc_ui_price_from_product(float(data.get("price_after_seller_discount") or 0))
            upsert_product(
                nm_id=data["nm_id"],
                brand=data.get("brand", "") or "",
                title=data.get("title", "") or "",
                price_before=float(data.get("price_before_discount") or 0),
                price_after=float(data.get("price_after_seller_discount") or 0),
                seller_id=data.get("seller_id"),
                seller_name=data.get("seller_name") or None,
                ui_price=int(ui_price) if ui_price is not None else None,
                table=table,
            )

        if rrc_auto is not None:
            set_rrc(data["nm_id"], rrc_auto, table=table)

        return jsonify({"ok": True, "item": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.post("/api/products/<int:nm_id>/rrc")
def post_rrc(nm_id: int):
    table = resolve_table()
    body = request.get_json(silent=True) or {}
    rrc = body.get("rrc", None)

    if rrc in ("", None):
        val = None
    else:
        try:
            val = float(rrc)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "rrc должно быть числом"}), 400

    try:
        set_rrc(nm_id, val, table=table)
        return jsonify({"ok": True, "nm_id": nm_id, "rrc": val})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.route("/api/products/<int:nm_id>/rrc", methods=["PATCH"])
def patch_rrc(nm_id: int):
    table = resolve_table()
    body = request.get_json(silent=True) or {}
    rrc = body.get("rrc", None)

    if rrc in ("", None):
        val = None
    else:
        try:
            val = float(rrc)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "rrc должно быть числом"}), 400

    try:
        set_rrc(nm_id, val, table=table)
        return jsonify({"ok": True, "nm_id": nm_id, "rrc": val})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.delete("/api/products/<int:nm_id>")
def delete_product(nm_id: int):
    table = resolve_table()
    try:
        db_delete(nm_id, table=table)
        return jsonify({"ok": True, "nm_id": nm_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.post("/api/products/refresh-batch")
def refresh_batch():
    """
    Пакетное обновление.
    Понимает ?table=products|products_ozon, silent=1, full=1, limit=...
    """
    try:
        table = resolve_table()
        limit = int(request.args.get("limit", BATCH_REFRESH_LIMIT))
        silent = bool(request.args.get("silent", type=int) or 0)
        full = bool(request.args.get("full", type=int) or 0)

        LOCK_NAME = f"{table}_refresh_full_lock"

        def do_one_batch(_limit: int):
            nm_ids = list_nm_ids_for_refresh(_limit, table=table)
            updated, errors = [], []

            for nm_id in nm_ids:
                try:
                    data = fetch_wb_price(nm_id)
                    price_before = float(data.get("price_before_discount") or 0)
                    price_after  = float(data.get("price_after_seller_discount") or 0)
                    if price_before <= 0 and price_after <= 0:
                        price_before, price_after = 0.0, -1.0

                    rrc_auto = calc_rrc_from_title(data.get("title"))
                    ui_price = calc_ui_price_from_product(price_after)

                    upsert_product(
                        nm_id=data["nm_id"],
                        brand=data.get("brand", "") or "",
                        title=data.get("title", "") or "",
                        price_before=price_before,
                        price_after=price_after,
                        seller_id=data.get("seller_id"),
                        seller_name=data.get("seller_name") or None,
                        ui_price=int(ui_price) if ui_price is not None else None,
                        table=table,
                    )
                    if rrc_auto is not None:
                        set_rrc(nm_id, rrc_auto, table=table)

                    updated.append(nm_id)
                except Exception as e:
                    errors.append({"nm_id": nm_id, "error": str(e)})

            remaining = count_needing_refresh(table=table)
            return updated, errors, remaining

        if not full:
            updated, errors, remaining = do_one_batch(limit)
            if not silent and remaining == 0 and is_work_time():
                try:
                    violators = list_sellers_with_violations(table=table)
                    for v in violators:
                        send_violation_alert(v["seller_id"], v.get("seller_name"))
                except Exception as e:
                    print(f"[alerts] send error: {e}")

            return jsonify({
                "ok": True,
                "requested": limit,
                "selected": len(updated) + len(errors),
                "updated_count": len(updated),
                "updated": updated,
                "errors_count": len(errors),
                "errors": errors,
                "remaining": remaining,
                "done": remaining == 0,
                "locked": False,
            })

        # full=1 под advisory-lock
        if not acquire_advisory_lock(LOCK_NAME, timeout=0):
            return jsonify({"ok": True, "locked": True}), 200

        try:
            total_updated = 0
            total_errors = 0
            last_errors = []
            safety_loops = 0
            MAX_LOOPS = 10000

            while True:
                safety_loops += 1
                if safety_loops > MAX_LOOPS:
                    break

                updated, errors, remaining = do_one_batch(limit)
                total_updated += len(updated)
                total_errors  += len(errors)
                last_errors = errors

                if remaining == 0:
                    break

            if not silent and is_work_time():
                try:
                    violators = list_sellers_with_violations(table=table)
                    for v in violators:
                        send_violation_alert(v["seller_id"], v.get("seller_name"))
                except Exception as e:
                    print(f"[alerts] send error: {e}")

            return jsonify({
                "ok": True,
                "mode": "full",
                "updated_total": total_updated,
                "errors_total": total_errors,
                "errors_last_batch": last_errors,
                "remaining": count_needing_refresh(table=table),
                "done": count_needing_refresh(table=table) == 0,
                "locked": False,
            })
        finally:
            release_advisory_lock(LOCK_NAME)

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.delete("/api/products")
def delete_all():
    """Удаляет все товары из выбранной таблицы."""
    try:
        table = resolve_table()
        removed = delete_all_products(table=table)
        return jsonify({"ok": True, "removed": removed})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.post("/api/upload-xlsx")
def upload_xlsx():
    """
    Импорт XLSX в выбранную таблицу (?table=...).
    Колонки nm/rrc определяются автоматически, как раньше.
    """
    try:
        table = resolve_table()
        file = request.files.get("file")
        if not file:
            return jsonify({"ok": False, "error": "Файл не передан"}), 400

        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        nm_col_arg = request.args.get("nm_col", type=int)
        rrc_col_arg = request.args.get("rrc_col", type=int)

        if nm_col_arg is not None or rrc_col_arg is not None:
            idx_nm, idx_rrc = nm_col_arg, rrc_col_arg
        else:
            idx_nm, idx_rrc = _detect_columns(ws)

        total = 0
        affected = 0
        skipped = 0
        errors = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1

            nm_val = None
            if idx_nm is not None and idx_nm < len(row):
                nm_val = _parse_nm_id(row[idx_nm])

            if not nm_val:
                skipped += 1
                continue

            rrc_val = None
            if idx_rrc is not None and idx_rrc < len(row):
                rrc_parsed = _parse_price_like(row[idx_rrc])
                if rrc_parsed is not None:
                    rrc_val = float(int(round(rrc_parsed)))

            try:
                upsert_product(
                    nm_id=nm_val,
                    brand="",
                    title="",
                    price_before=0.0,
                    price_after=0.0,
                    seller_id=None,
                    seller_name=None,
                    ui_price=None,
                    table=table,
                )
                if rrc_val is not None:
                    set_rrc(nm_val, rrc_val, table=table)
                affected += 1
            except Exception:
                errors += 1

        return jsonify({
            "ok": True,
            "total": total,
            "affected": affected,
            "skipped": skipped,
            "errors_count": errors,
            "columns": {
                "nm_idx": idx_nm,
                "rrc_idx": idx_rrc,
                "header_nm": (ws.cell(row=1, column=(idx_nm+1)).value if idx_nm is not None else None),
                "header_rrc": (ws.cell(row=1, column=(idx_rrc+1)).value if idx_rrc is not None else None),
            }
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.post("/api/upload-ozon-csv")
def upload_ozon_csv():
    # какую таблицу заполняем: по умолчанию Ozon
    table = DB_TABLE_OZON

    file = request.files.get("file") or request.files.get("csv")
    if not file:
        return jsonify({"ok": False, "error": "Файл не передан"}), 400

    data_bytes = file.read()
    total = affected = errors = 0

    for rec in iter_ozon_csv_rows(data_bytes):
        total += 1
        try:
            # Записываем всё, что пришло из CSV. Продавца тоже сохраняем.
            upsert_product(
                nm_id=rec["nm_id"],
                brand="",                                     # в CSV бренда нет — оставляем пусто
                title=rec.get("title") or "",
                price_before=rec.get("price_before") or 0.0,
                price_after=rec.get("price_after") or 0.0,
                seller_id=rec.get("seller_id"),
                seller_name=rec.get("seller_name"),
                ui_price=None,
                table=table,
            )
            # Если захочешь авто-РРЦ — раскомментируй:
            rrc_auto = calc_rrc_from_title(rec.get("title"))
            if rrc_auto is not None: set_rrc(rec["nm_id"], rrc_auto, table=table)

            affected += 1
        except Exception:
            errors += 1

    return jsonify({
        "ok": True,
        "source": "csv",
        "total": total,
        "affected": affected,
        "errors_count": errors,
        "table": table,
    })

@application.get("/api/sellers/<int:seller_id>/violations")
def seller_violations(seller_id: int):
    """Список nm_id с нарушением для селлера в выбранной таблице (?table=...)."""
    try:
        table = resolve_table()
        rows = list_violations_for_seller(seller_id, table=table)
        return jsonify({"items": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.get("/api/stats")
def stats():
    """Сводка по выбранной таблице (?table=...)."""
    try:
        table = resolve_table()
        total_rows = count_all_rows(table=table)
        viol_count = count_violations(table=table)
        remaining = count_needing_refresh(table=table)
        return jsonify({
            "ok": True,
            "total_rows": total_rows,
            "violations": viol_count,
            "needing_refresh": remaining,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@application.get("/")
def index_root():
    return "API OK with DB. try /api/health"